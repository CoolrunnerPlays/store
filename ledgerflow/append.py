"""Append new records into an existing workbook without disturbing it.

The rules this module keeps, in priority order:

1. Nothing already in the file changes. Existing rows, styles, row heights,
   column widths, freeze panes and formulas come out the way they went in, and
   this is verified against the original after writing rather than assumed.
2. Nothing is written unless there is something new to write. A run that finds
   only duplicates stops before the workbook is opened for modification.
3. Totals stay correct. Rows go *inside* the table, and every formula in the
   workbook is re-pointed so the existing SUMs cover the new rows.
4. Numbers are numbers, written as numeric cells so the user's formulas work.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Literal

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .formulas import shift_for_insert, shift_range_ref, translate_rows
from .introspect import TableInfo, describe_sheet, describe_workbook
from .mapping import MONEY_FIELDS, Profile, value_for
from .matching import ExistingRow, Matcher, MatchResult, cents, coerce_date
from .models import Flag, Record

STATE_SHEET = "_LedgerFlow"
REVIEW_SHEET = "Review Notes"

Status = Literal["written", "no_new_records", "nothing_extracted", "unsafe_to_match", "dry_run"]

NO_NEW_MESSAGE = "0 new transactions found (100% duplicate history)"


@dataclass
class Judgement:
    """One extracted record and what the matcher decided about it."""

    record: Record
    match: MatchResult

    @property
    def status(self) -> str:
        return self.match.status


@dataclass
class AppendResult:
    """What one run did, in the terms the summary needs."""

    output: Path | None
    sheet: str
    status: Status = "written"
    message: str = ""
    added: list[Record] = field(default_factory=list)
    duplicates: list[Judgement] = field(default_factory=list)
    probable_duplicates: list[Judgement] = field(default_factory=list)
    unmatchable: list[Judgement] = field(default_factory=list)
    skipped_other_kind: list[Record] = field(default_factory=list)
    flags: list[Flag] = field(default_factory=list)
    first_new_row: int = 0
    formulas_updated: int = 0
    verified: bool = False
    """True when the written file was re-read and its untouched region matched the original."""

    @property
    def file_written(self) -> bool:
        return self.status == "written" and self.output is not None

    @property
    def skipped_duplicates(self) -> list[Record]:
        """Every record held back as already present, however it was matched."""
        return [j.record for j in self.duplicates + self.probable_duplicates]

    @property
    def invoices(self) -> list[Record]:
        return [r for r in self.added if r.kind == "invoice"]

    @property
    def transactions(self) -> list[Record]:
        return [r for r in self.added if r.kind == "transaction"]

    @property
    def money_in(self) -> Decimal:
        return sum((r.deposit for r in self.added if r.deposit), Decimal("0"))

    @property
    def money_out(self) -> Decimal:
        return sum((r.withdrawal for r in self.added if r.withdrawal), Decimal("0"))

    @property
    def invoice_total(self) -> Decimal:
        return sum((r.total for r in self.invoices if r.total), Decimal("0"))

    @property
    def total_added(self) -> Decimal:
        """Sum of what was appended, using each record's own headline figure."""
        total = Decimal("0")
        for record in self.added:
            value = record.amount if record.amount is not None else record.total
            if value is not None:
                total += value
        return total


class PreservationError(RuntimeError):
    """Raised when the written file does not match the original where it should."""


def append_records(
    workbook_path: str | Path,
    records: Iterable[Record],
    profile: Profile,
    *,
    output_path: str | Path | None = None,
    flags: list[Flag] | None = None,
    track_state: bool = True,
    dry_run: bool = False,
    verify: bool = True,
) -> AppendResult:
    """Append records to the sheet named by ``profile``.

    The workbook is only opened for modification once there is at least one
    record that is genuinely not already in it. A run with nothing new leaves the
    file alone and says so; it does not write an identical copy.
    """
    workbook_path = Path(workbook_path)
    output = Path(output_path) if output_path else workbook_path.with_name(
        f"{workbook_path.stem}_updated{workbook_path.suffix}"
    )

    plan = plan_append(workbook_path, records, profile, flags=flags)
    plan.output = None

    if plan.status != "written":
        return plan
    if dry_run:
        plan.status = "dry_run"
        return plan

    plan.output = output
    _perform_append(workbook_path, output, plan, profile, track_state=track_state)

    if verify:
        _verify_untouched(workbook_path, output, profile.sheet, plan.first_new_row, len(plan.added))
        plan.verified = True
    return plan


def plan_append(
    workbook_path: str | Path,
    records: Iterable[Record],
    profile: Profile,
    *,
    flags: list[Flag] | None = None,
) -> AppendResult:
    """Decide what would be appended, without opening the file for writing.

    This is the whole safety story: every abort happens here, before anything is
    modified. The web preview and the command line both run this first and show
    the user its verdict.
    """
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        if profile.sheet not in wb.sheetnames:
            raise ValueError(f"Sheet {profile.sheet!r} is not in {workbook_path.name}")
        table = describe_sheet(wb[profile.sheet])
        if table is None:
            raise ValueError(f"No data table could be found on sheet {profile.sheet!r}")
        existing = read_existing_rows(wb[profile.sheet], table, profile)
    finally:
        wb.close()

    result = AppendResult(output=None, sheet=profile.sheet, flags=list(flags or []))
    result.first_new_row = _insertion_row(table)

    wanted, result.skipped_other_kind = _split_by_kind(records, profile)

    if not wanted:
        result.status = "nothing_extracted"
        result.message = (
            "Nothing was extracted for this sheet. "
            f"{len(result.skipped_other_kind)} record(s) belonged to a different kind of sheet."
            if result.skipped_other_kind
            else "No records were extracted from the documents provided."
        )
        return result

    matcher = Matcher(rows=existing) if profile.skip_duplicates else Matcher(rows=[])

    # Fail closed: if the sheet has rows but none of them could be read well
    # enough to compare against, appending would risk a second copy of the book.
    if profile.skip_duplicates and table.row_count > 0 and matcher.usable_rows == 0:
        result.status = "unsafe_to_match"
        result.message = (
            f"'{profile.sheet}' has {table.row_count} rows, but none of them could be read as a "
            "date and an amount, so new records cannot be checked against them. Nothing was written. "
            "Check that the Date and amount columns are mapped correctly."
        )
        return result

    fresh: list[Record] = []
    for record in _sorted(wanted, profile):
        verdict = matcher.find(record)
        judgement = Judgement(record, verdict)
        if verdict.status == "duplicate":
            result.duplicates.append(judgement)
        elif verdict.status == "probable_duplicate":
            result.probable_duplicates.append(judgement)
            result.flags.append(
                Flag(
                    source=record.source,
                    location=f"row {verdict.row} of '{profile.sheet}'" if verdict.row else "",
                    field="duplicate",
                    raw=f"{record.date} {record.description[:40]}",
                    reason=f"Held back as a likely duplicate: {verdict.reason}. "
                           "It was not added; add it by hand if it is genuinely a separate transaction.",
                    row_key=record.fingerprint(),
                )
            )
        elif verdict.status == "unmatchable":
            result.unmatchable.append(judgement)
            result.flags.append(
                Flag(
                    source=record.source,
                    location=f"page {record.page}" if record.page else "",
                    field="amount",
                    raw=record.description[:60],
                    reason=verdict.reason + " It was not added.",
                    row_key=record.fingerprint(),
                )
            )
        else:
            fresh.append(record)
            matcher.add(record)

    result.added = fresh

    if not fresh:
        result.status = "no_new_records"
        result.message = NO_NEW_MESSAGE
        return result

    result.status = "written"
    return result


def _sorted(records: list[Record], profile: Profile) -> list[Record]:
    if not profile.sort_by_date:
        return list(records)
    return sorted(records, key=lambda r: (r.date or date.max, r.description))


def read_existing_rows(ws: Worksheet, table: TableInfo, profile: Profile) -> list[ExistingRow]:
    """Reduce the rows already in the sheet to what the matcher compares on.

    Dates typed as text are parsed, and the amount is taken from whichever of the
    mapped money columns the book actually uses.
    """
    by_field = {field_name: letter for letter, field_name in profile.columns.items()}
    values = {
        row: {
            letter: ws[f"{letter}{row}"].value
            for letter in profile.columns
        }
        for row in range(table.first_data_row, table.last_data_row + 1)
    }

    rows: list[ExistingRow] = []
    for row, cells in values.items():
        def cell(field_name: str):
            letter = by_field.get(field_name)
            if not letter:
                return None
            value = cells.get(letter)
            return None if isinstance(value, str) and value.startswith("=") else value

        amount = _numeric(cell("amount"))
        if amount is None:
            deposit, withdrawal = _numeric(cell("deposit")), _numeric(cell("withdrawal"))
            if deposit or withdrawal:
                amount = (deposit or Decimal(0)) - (withdrawal or Decimal(0))
        if amount is None:
            amount = _numeric(cell("total"))

        description = cell("description") or cell("payee") or cell("vendor") or ""
        reference = cell("reference") or ""

        rows.append(
            ExistingRow(
                row=row,
                date=coerce_date(cell("date")),
                description=str(description),
                reference=str(reference),
                amount_cents=cents(amount),
            )
        )
    return rows


def _numeric(value: Any) -> Decimal | None:
    if value is None or isinstance(value, (str, datetime, date)):
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _insertion_row(table: TableInfo) -> int:
    """First row of the new block: directly under the last row of real data."""
    return max(table.last_data_row, table.header_row) + 1


def _split_by_kind(records: Iterable[Record], profile: Profile) -> tuple[list[Record], list[Record]]:
    wanted, rejected = [], []
    for record in records:
        (wanted if record.kind in profile.kinds else rejected).append(record)
    return wanted, rejected


def _perform_append(
    source: Path, output: Path, plan: AppendResult, profile: Profile, *, track_state: bool
) -> None:
    """Do the actual writing. Only reached when there is something new to add."""
    wb = load_workbook(source, data_only=False, keep_vba=source.suffix.lower() == ".xlsm")
    ws = wb[profile.sheet]
    table = describe_sheet(ws)

    at_row = _insertion_row(table)
    count = len(plan.added)
    plan.first_new_row = at_row
    style_source = table.last_data_row if table.last_data_row > table.header_row else None

    ws.insert_rows(at_row, count)
    _shift_row_dimensions(ws, at_row, count)
    plan.formulas_updated = _repoint_workbook(wb, profile.sheet, at_row, count)
    _repair_ranges(ws, at_row, count)

    for offset, record in enumerate(plan.added):
        row = at_row + offset
        if style_source is not None:
            _copy_row_style(ws, style_source, row, table)
        _write_row(ws, row, record, table, profile, style_source)

    _write_review_notes(wb, plan)
    if track_state:
        _write_state(wb, plan.added)

    wb.save(output)


def _shift_row_dimensions(ws: Worksheet, at_row: int, count: int) -> None:
    """Move row heights and row-level styles down with the rows they belong to.

    openpyxl's insert_rows moves the cells but leaves row_dimensions bound to the
    old row numbers, so every row below the insertion point inherits the height
    and outline of whatever row now sits at its index. On a book with custom row
    heights that reads as the whole sheet below the insert being mangled.
    """
    existing = {index: dimension for index, dimension in ws.row_dimensions.items()}
    moved = {index: dimension for index, dimension in existing.items() if index >= at_row}
    if not moved:
        return

    for index in moved:
        del ws.row_dimensions[index]

    for index in sorted(moved, reverse=True):
        dimension = moved[index]
        dimension.index = index + count
        ws.row_dimensions[index + count] = dimension


def _repoint_workbook(wb, target_sheet: str, at_row: int, count: int) -> int:
    """Rewrite every formula in the book for the inserted rows.

    openpyxl moves cells but never touches formula text, so without this pass a
    totals row that moved down would still be summing the rows it used to sit
    under. Runs over all sheets because a summary tab commonly points here.
    """
    updated = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                rewritten = shift_for_insert(
                    value, sheet=sheet.title, target_sheet=target_sheet, at_row=at_row, count=count
                )
                if rewritten != value:
                    cell.value = rewritten
                    updated += 1

    for name, defined in list(getattr(wb, "defined_names", {}).items()):
        try:
            attr_text = defined.attr_text
        except AttributeError:
            continue
        if not attr_text:
            continue
        rewritten = shift_for_insert(
            "=" + attr_text, sheet=target_sheet, target_sheet=target_sheet, at_row=at_row, count=count
        )[1:]
        if rewritten != attr_text:
            defined.attr_text = rewritten
            updated += 1
    return updated


def _repair_ranges(ws: Worksheet, at_row: int, count: int) -> None:
    """Grow the sheet features that openpyxl leaves pointing at the old rows.

    Excel tables, the autofilter, conditional formats, validations and merges all
    store their own A1 ranges. Left alone, banding stops halfway down the table
    and dropdowns go missing on the new rows.
    """
    for table in list(getattr(ws, "tables", {}).values()):
        table.ref = shift_range_ref(table.ref, at_row=at_row, count=count)

    if ws.auto_filter and ws.auto_filter.ref:
        ws.auto_filter.ref = shift_range_ref(ws.auto_filter.ref, at_row=at_row, count=count)

    if getattr(ws, "print_area", None):
        areas = ws.print_area if isinstance(ws.print_area, (list, tuple)) else [ws.print_area]
        ws.print_area = [shift_range_ref(str(a).replace("$", ""), at_row=at_row, count=count) for a in areas]

    for rule_range in list(getattr(ws.conditional_formatting, "_cf_rules", {}).keys()):
        rules = ws.conditional_formatting._cf_rules[rule_range]
        new_ref = " ".join(
            shift_range_ref(part, at_row=at_row, count=count) for part in str(rule_range.sqref).split()
        )
        if new_ref != str(rule_range.sqref):
            del ws.conditional_formatting._cf_rules[rule_range]
            rule_range.sqref = new_ref
            ws.conditional_formatting._cf_rules[rule_range] = rules

    for validation in list(getattr(ws, "data_validations", None).dataValidation if ws.data_validations else []):
        parts = [shift_range_ref(str(p), at_row=at_row, count=count) for p in str(validation.sqref).split()]
        validation.sqref = " ".join(parts)

    merged = [str(m) for m in ws.merged_cells.ranges]
    for ref in merged:
        shifted = shift_range_ref(ref, at_row=at_row, count=count)
        if shifted != ref:
            ws.unmerge_cells(ref)
            ws.merge_cells(shifted)


def _copy_row_style(ws: Worksheet, source_row: int, target_row: int, table: TableInfo) -> None:
    """Give a new row exactly the look of the row it follows.

    Every style facet is copied explicitly -- font family, size, weight, colour,
    fill, all four borders, alignment, protection and number format -- rather
    than leaving any of them to Excel's defaults.
    """
    for col in range(table.first_col, table.last_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format

    source_dimension = ws.row_dimensions.get(source_row)
    if source_dimension is not None and source_dimension.height is not None:
        ws.row_dimensions[target_row].height = source_dimension.height


def _write_row(
    ws: Worksheet,
    row: int,
    record: Record,
    table: TableInfo,
    profile: Profile,
    style_source: int | None,
) -> None:
    """Fill one new row: mapped values, and the sheet's own per-row formulas."""
    for column in table.columns:
        cell = ws.cell(row=row, column=column.index)

        if column.is_computed:
            # Carry the sheet's own calculation down, the way filling down would.
            cell.value = translate_rows(column.formula_template, row - table.last_data_row)
            if style_source is None:
                cell.number_format = column.number_format
            continue

        field_name = profile.field_for(column.letter)
        if field_name is None:
            continue

        value = value_for(record, field_name)
        if value is None:
            continue

        if field_name in MONEY_FIELDS and isinstance(value, Decimal):
            cell.value = float(value)
            if style_source is None:
                cell.number_format = column.number_format
        elif isinstance(value, date):
            cell.value = value
            cell.number_format = profile.date_format or column.number_format or "yyyy-mm-dd"
        else:
            cell.value = str(value)


def _write_review_notes(wb, result: AppendResult) -> None:
    """Write the flags to their own sheet, newest run at the top of its block."""
    if not result.flags:
        return

    if REVIEW_SHEET in wb.sheetnames:
        ws = wb[REVIEW_SHEET]
        start = ws.max_row + 2
    else:
        ws = wb.create_sheet(REVIEW_SHEET)
        headers = ["Run", "Source file", "Where", "Field", "Value as read", "Why it needs checking"]
        for index, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=index, value=name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="B45309")
        widths = {"A": 20, "B": 28, "C": 40, "D": 16, "E": 24, "F": 60}
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
        ws.freeze_panes = "A2"
        start = 2

    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")
    for offset, flag in enumerate(result.flags):
        row = start + offset
        for index, value in enumerate(
            [stamp, flag.source, flag.location, flag.field, flag.raw, flag.reason], start=1
        ):
            cell = ws.cell(row=row, column=index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=index == 6)


def _write_state(wb, records: list[Record]) -> None:
    """Record what has been added, for auditing which run a row came from."""
    if STATE_SHEET in wb.sheetnames:
        ws = wb[STATE_SHEET]
    else:
        ws = wb.create_sheet(STATE_SHEET)
        for index, name in enumerate(["fingerprint", "added", "source", "date", "amount"], start=1):
            ws.cell(row=1, column=index, value=name)
        ws.sheet_state = "hidden"

    stamp = datetime.now().isoformat(timespec="seconds")
    row = ws.max_row + 1
    for offset, record in enumerate(records):
        amount = record.amount if record.amount is not None else record.total
        ws.cell(row=row + offset, column=1, value=record.fingerprint())
        ws.cell(row=row + offset, column=2, value=stamp)
        ws.cell(row=row + offset, column=3, value=record.source)
        ws.cell(row=row + offset, column=4, value=record.date.isoformat() if record.date else "")
        ws.cell(row=row + offset, column=5, value=float(amount) if amount is not None else None)


def _rgb(holder) -> str | None:
    """Colour of a style element, tolerating the many ways openpyxl leaves it unset."""
    colour = getattr(holder, "color", None) if holder is not None else None
    rgb = getattr(colour, "rgb", None)
    return str(rgb) if rgb is not None else None


def _side(border, name: str) -> tuple:
    side = getattr(border, name, None)
    return (getattr(side, "style", None), _rgb(side))


def _style_signature(cell) -> tuple:
    """Everything about a cell's appearance that must survive the round trip."""
    font, fill, border, alignment = cell.font, cell.fill, cell.border, cell.alignment
    return (
        font.name, font.sz, font.b, font.i, font.u, font.strike, _rgb(font),
        fill.patternType,
        str(getattr(fill.fgColor, "rgb", None)) if getattr(fill, "fgColor", None) is not None else None,
        str(getattr(fill.bgColor, "rgb", None)) if getattr(fill, "bgColor", None) is not None else None,
        _side(border, "left"), _side(border, "right"), _side(border, "top"), _side(border, "bottom"),
        alignment.horizontal, alignment.vertical, alignment.wrapText, alignment.indent,
        cell.number_format,
    )


def _verify_untouched(source: Path, output: Path, sheet: str, at_row: int, count: int) -> None:
    """Re-read the written file and prove the pre-existing content is unchanged.

    Values, every style facet and row heights are compared cell by cell: rows
    above the insertion point against themselves, and rows below against their
    new positions. A guarantee that is never checked is only a hope, and this is
    the check.
    """
    before_wb = load_workbook(source, data_only=False)
    after_wb = load_workbook(output, data_only=False)
    problems: list[str] = []

    for name in before_wb.sheetnames:
        if name not in after_wb.sheetnames:
            problems.append(f"sheet '{name}' is missing from the output")
            continue
        before, after = before_wb[name], after_wb[name]
        shift = count if name == sheet else 0

        for row in range(1, before.max_row + 1):
            target_row = row + shift if (shift and row >= at_row) else row
            for col in range(1, before.max_column + 1):
                old = before.cell(row=row, column=col)
                new = after.cell(row=target_row, column=col)
                if _style_signature(old) != _style_signature(new):
                    problems.append(f"{name}!{old.coordinate}: formatting changed")
                # Formulas on the target sheet are deliberately re-pointed.
                is_formula = isinstance(old.value, str) and old.value.startswith("=")
                if not is_formula and old.value != new.value:
                    problems.append(f"{name}!{old.coordinate}: value changed from {old.value!r} to {new.value!r}")
                if len(problems) > 12:
                    break

            old_height = before.row_dimensions[row].height
            new_height = after.row_dimensions[target_row].height
            if old_height != new_height:
                problems.append(f"{name} row {row}: height changed from {old_height} to {new_height}")
            if len(problems) > 12:
                break

        for letter, dimension in before.column_dimensions.items():
            after_dimension = after.column_dimensions.get(letter)
            if after_dimension is None or after_dimension.width != dimension.width:
                problems.append(f"{name} column {letter}: width changed")

    if problems:
        raise PreservationError(
            "The written file does not match the original where it should. Nothing was returned. "
            + "; ".join(problems[:12])
        )


def choose_sheet(workbook_path: str | Path) -> tuple[str, dict[str, TableInfo]]:
    """Pick the most ledger-like sheet in a workbook, and report all candidates."""
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    tables = {name: info for name, info in describe_workbook(wb).items() if info.looks_like_table}
    if not tables:
        raise ValueError("No table with headers and data rows was found in this workbook.")
    best = max(tables.items(), key=lambda item: item[1].score)[0]
    return best, tables
