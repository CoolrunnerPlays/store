"""Work out the shape of a workbook the user already keeps by hand.

The point of this module is that nobody has to describe their spreadsheet to
the tool. It opens the file, finds the header row, the last row of real data,
the columns that carry per-row formulas and the totals that sit underneath, and
reports all of it so the append step knows exactly where to write and what to
extend.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

MAX_HEADER_SCAN = 30
"""How far down to look for a header row before giving up."""

TOTAL_WORDS = re.compile(r"\b(total|totals|sum|subtotal|balance c/?f|grand total|net)\b", re.IGNORECASE)


@dataclass
class ColumnInfo:
    """One column of the detected table."""

    index: int
    letter: str
    header: str
    formula_template: str | None = None
    """Formula found on the last data row, if this column is computed per-row."""
    number_format: str = "General"
    sample_values: list[Any] = field(default_factory=list)
    inferred_type: str = "text"
    """One of text, number, money, date."""

    @property
    def is_computed(self) -> bool:
        return self.formula_template is not None


@dataclass
class TableInfo:
    """The detected data table on one worksheet."""

    sheet: str
    header_row: int
    first_col: int
    last_col: int
    first_data_row: int
    last_data_row: int
    columns: list[ColumnInfo]
    total_rows: list[int] = field(default_factory=list)
    """Rows below the data that hold totals or other aggregates."""
    excel_table: str | None = None
    """Name of the Excel ListObject covering this range, if any."""
    row_count: int = 0

    @property
    def score(self) -> float:
        """How much this looks like the ledger the user actually maintains.

        Used to pick a default sheet when a workbook has several. A wide table
        with many rows and named headers beats a two-cell summary block.
        """
        named = sum(1 for c in self.columns if c.header)
        if self.row_count < 1 or named < 3:
            return 0.0
        return named * 2 + min(self.row_count, 200) * 0.5

    @property
    def looks_like_table(self) -> bool:
        return self.score > 0

    def column_by_header(self, header: str) -> ColumnInfo | None:
        target = _norm_header(header)
        for column in self.columns:
            if _norm_header(column.header) == target:
                return column
        return None


def _norm_header(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(text).lower())


def _cell_is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _looks_like_label(value: Any) -> bool:
    """A header cell is short text that is not itself a date or a number."""
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text or len(text) > 60:
        return False
    return not re.fullmatch(r"[\d\s.,$£€%()+-]+", text)


def find_header_row(ws: Worksheet) -> int | None:
    """Return the 1-based row that most plausibly holds the column headers.

    Scores each of the first rows by how many label-shaped cells it has, and
    requires the row underneath to carry something, so a stray title line above
    an empty region is not mistaken for the header.
    """
    best_row, best_score = None, 0
    limit = min(ws.max_row, MAX_HEADER_SCAN)
    for row in range(1, limit + 1):
        values = [c.value for c in ws[row]]
        score = sum(1 for v in values if _looks_like_label(v))
        if score < 2 or score <= best_score:
            continue
        below = [c.value for c in ws[row + 1]] if row + 1 <= ws.max_row else []
        if not any(not _cell_is_blank(v) for v in below):
            continue
        best_row, best_score = row, score
    return best_row


def _column_span(ws: Worksheet, header_row: int) -> tuple[int, int]:
    """Widest run of populated header cells, tolerating a single blank spacer."""
    populated = [c.column for c in ws[header_row] if not _cell_is_blank(c.value)]
    if not populated:
        return 1, 1
    return min(populated), max(populated)


def _is_total_row(ws: Worksheet, row: int, first_col: int, last_col: int) -> bool:
    """A row is a totals row when it holds an aggregate formula or says so in words."""
    for col in range(first_col, last_col + 1):
        value = ws.cell(row=row, column=col).value
        if isinstance(value, str):
            if value.startswith("=") and re.search(r"\b(SUM|SUBTOTAL|SUMIF|SUMIFS|COUNT|AVERAGE)\b", value, re.I):
                return True
            if TOTAL_WORDS.search(value) and len(value) < 40:
                return True
    return False


def _last_data_row(ws: Worksheet, header_row: int, first_col: int, last_col: int) -> tuple[int, list[int]]:
    """Find the final row of real data and the totals rows that follow it.

    Walks up from the bottom of the used range so trailing blank rows and a
    totals block at the foot of the table do not get counted as data.
    """
    totals: list[int] = []
    last = header_row
    for row in range(ws.max_row, header_row, -1):
        values = [ws.cell(row=row, column=c).value for c in range(first_col, last_col + 1)]
        if all(_cell_is_blank(v) for v in values):
            continue
        if _is_total_row(ws, row, first_col, last_col):
            totals.append(row)
            continue
        last = row
        break
    return last, sorted(totals)


def _infer_type(values: list[Any], number_format: str) -> str:
    for value in values:
        if isinstance(value, (datetime, date)):
            return "date"
    if any(sym in number_format for sym in ("$", "£", "€", "#,##0.00")):
        return "money"
    if any(isinstance(v, (int, float)) for v in values):
        return "number"
    return "text"


def describe_sheet(ws: Worksheet) -> TableInfo | None:
    """Detect the table on one worksheet, or None when the sheet holds no table."""
    header_row = find_header_row(ws)
    if header_row is None:
        return None
    first_col, last_col = _column_span(ws, header_row)
    last_data, totals = _last_data_row(ws, header_row, first_col, last_col)
    first_data = header_row + 1

    columns: list[ColumnInfo] = []
    for col in range(first_col, last_col + 1):
        header = ws.cell(row=header_row, column=col).value
        header_text = str(header).strip() if header is not None else ""
        samples: list[Any] = []
        for row in range(first_data, min(last_data, first_data + 25) + 1):
            value = ws.cell(row=row, column=col).value
            if not _cell_is_blank(value):
                samples.append(value)

        formula = None
        if last_data >= first_data:
            bottom = ws.cell(row=last_data, column=col).value
            if isinstance(bottom, str) and bottom.startswith("="):
                formula = bottom

        cell = ws.cell(row=max(first_data, min(last_data, first_data)), column=col)
        number_format = cell.number_format or "General"
        columns.append(
            ColumnInfo(
                index=col,
                letter=get_column_letter(col),
                header=header_text,
                formula_template=formula,
                number_format=number_format,
                sample_values=samples[:5],
                inferred_type=_infer_type(samples, number_format),
            )
        )

    table_name = None
    for name, table in getattr(ws, "tables", {}).items():
        table_name = name
        break

    return TableInfo(
        sheet=ws.title,
        header_row=header_row,
        first_col=first_col,
        last_col=last_col,
        first_data_row=first_data,
        last_data_row=last_data,
        columns=columns,
        total_rows=totals,
        excel_table=table_name,
        row_count=max(0, last_data - first_data + 1),
    )


def describe_workbook(wb) -> dict[str, TableInfo]:
    """Detect a table on every worksheet that has one, keyed by sheet name."""
    found: dict[str, TableInfo] = {}
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        info = describe_sheet(ws)
        if info is not None:
            found[ws.title] = info
    return found
