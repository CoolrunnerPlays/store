"""Builders for workbooks that stand in for the user's own filled spreadsheet.

Each builder deliberately includes the things that make appending hard: a title
above the header, per-row formulas, a totals row with SUM ranges, currency
formats, an Excel table, and a summary sheet that points at the data sheet.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

MONEY = '"$"#,##0.00'
DATEFMT = "yyyy-mm-dd"


def _style_header(ws, row: int, first: int, last: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="B0B0B0")
    for col in range(first, last + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=thin)


def build_ledger(path, *, rows: int = 8, with_table: bool = False):
    """A bank-ledger workbook: title, headers, data, per-row Net formula, totals."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws["A1"] = "Weekly Cash Ledger 2026"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Date", "Description", "Category", "Deposit", "Withdrawal", "Net", "Reference"]
    for idx, name in enumerate(headers, start=1):
        ws.cell(row=3, column=idx, value=name)
    _style_header(ws, 3, 1, len(headers))

    start = 4
    for i in range(rows):
        row = start + i
        ws.cell(row=row, column=1, value=date(2026, 1, 5 + i)).number_format = DATEFMT
        ws.cell(row=row, column=2, value=f"EXISTING PAYEE {i + 1}")
        ws.cell(row=row, column=3, value="Operations")
        ws.cell(row=row, column=4, value=100 + i).number_format = MONEY
        ws.cell(row=row, column=5, value=40 + i).number_format = MONEY
        ws.cell(row=row, column=6, value=f"=D{row}-E{row}").number_format = MONEY
        ws.cell(row=row, column=7, value=f"REF{1000 + i}")

    last = start + rows - 1
    total_row = last + 2
    ws.cell(row=total_row, column=2, value="TOTAL")
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    for col in ("D", "E", "F"):
        cell = ws[f"{col}{total_row}"]
        cell.value = f"=SUM({col}{start}:{col}{last})"
        cell.number_format = MONEY
        cell.font = Font(bold=True)

    ws.column_dimensions["B"].width = 32
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{last}"

    if with_table:
        table = Table(displayName="LedgerTable", ref=f"A3:G{last}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(table)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Deposits"
    summary["B1"] = f"=SUM(Transactions!D{start}:D{last})"
    summary["A2"] = "Withdrawals"
    summary["B2"] = f"=SUM(Transactions!E{start}:E{last})"
    summary["A3"] = "Net"
    summary["B3"] = "=B1-B2"

    wb.save(path)
    return path


def build_invoice_book(path, *, rows: int = 4):
    """An invoice register with a computed Total column and a SUM footer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = ["Invoice Date", "Vendor", "Invoice No", "Description", "Subtotal", "Tax", "Total"]
    for idx, name in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=name)
    _style_header(ws, 1, 1, len(headers))

    start = 2
    for i in range(rows):
        row = start + i
        ws.cell(row=row, column=1, value=date(2026, 2, 3 + i)).number_format = DATEFMT
        ws.cell(row=row, column=2, value=f"Vendor {i + 1}")
        ws.cell(row=row, column=3, value=f"INV-{500 + i}")
        ws.cell(row=row, column=4, value="Consulting services")
        ws.cell(row=row, column=5, value=1000 + i * 10).number_format = MONEY
        ws.cell(row=row, column=6, value=(1000 + i * 10) * 0.1).number_format = MONEY
        ws.cell(row=row, column=7, value=f"=E{row}+F{row}").number_format = MONEY

    last = start + rows - 1
    total_row = last + 1
    ws.cell(row=total_row, column=4, value="TOTAL")
    for col in ("E", "F", "G"):
        ws[f"{col}{total_row}"] = f"=SUM({col}{start}:{col}{last})"
        ws[f"{col}{total_row}"].number_format = MONEY

    wb.save(path)
    return path


def build_populated_ledger(path, records, *, variant: str = "identical"):
    """A book a person already filled in, holding exactly ``records``.

    Reproduces the audit run: the workbook is complete before LedgerFlow ever
    sees it, so a correct run must append nothing. ``variant`` chooses how far
    the human's typing drifts from what the parser reads out of the PDF.

    * identical  - the same text the parser produces
    * reworded   - the payee typed a person's own way
    * textdates  - dates stored as text, which is very common by hand
    * signflip   - outgoings recorded as positive in a single Amount column
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws["A1"] = "Cash Ledger"
    ws["A1"].font = Font(bold=True, size=14, name="Calibri")

    single_amount = variant == "signflip"
    headers = (["Date", "Description", "Category", "Amount", "Reference"] if single_amount
               else ["Date", "Description", "Category", "Deposit", "Withdrawal", "Net", "Reference"])
    for index, name in enumerate(headers, start=1):
        ws.cell(row=3, column=index, value=name)
    _style_header(ws, 3, 1, len(headers))

    start = 4
    row = start
    for record in records:
        description = record.description
        if variant == "reworded":
            description = (description.title() if "PAYROLL" in description
                           else description.replace(" ", " - ", 1))

        ws.cell(row=row, column=1,
                value=record.date.strftime("%m/%d/%Y") if variant == "textdates" else record.date)
        if variant != "textdates":
            ws.cell(row=row, column=1).number_format = DATEFMT
        ws.cell(row=row, column=2, value=description).font = Font(name="Georgia", size=10)
        ws.cell(row=row, column=3, value="Operations")

        if single_amount:
            amount = record.amount or Decimal(0)
            ws.cell(row=row, column=4, value=float(abs(amount))).number_format = MONEY
            ws.cell(row=row, column=5, value=record.reference or None)
        else:
            if record.deposit:
                ws.cell(row=row, column=4, value=float(record.deposit)).number_format = MONEY
            if record.withdrawal:
                ws.cell(row=row, column=5, value=float(record.withdrawal)).number_format = MONEY
            ws.cell(row=row, column=6, value=f"=D{row}-E{row}").number_format = MONEY
            ws.cell(row=row, column=7, value=record.reference or None)

        # Alternating custom heights: openpyxl does not move these on insert.
        ws.row_dimensions[row].height = 22 if row % 2 == 0 else 15
        row += 1

    last = row - 1
    total_row = last + 2
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True, name="Georgia")
    ws.row_dimensions[total_row].height = 34
    for column in (["D"] if single_amount else ["D", "E", "F"]):
        cell = ws[f"{column}{total_row}"]
        cell.value = f"=SUM({column}{start}:{column}{last})"
        cell.number_format = MONEY
        cell.font = Font(bold=True)

    wb.save(path)
    return path
