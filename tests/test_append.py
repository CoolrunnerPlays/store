"""End-to-end checks that the user's own workbook survives an append intact."""

from __future__ import annotations

import re
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from fixtures import build_invoice_book, build_ledger
from ledgerflow.append import append_records, choose_sheet
from ledgerflow.extract import extract_records
from ledgerflow.introspect import describe_sheet
from ledgerflow.mapping import suggest_profile


def sum_of(ws, ref: str) -> Decimal:
    """Add up a range the way Excel would, over the literal values present."""
    total = Decimal("0")
    for row in ws[ref]:
        for cell in row:
            if isinstance(cell.value, (int, float)):
                total += Decimal(str(cell.value))
    return total


def range_in(formula: str) -> str:
    return re.search(r"SUM\((.+?)\)", formula).group(1)


@pytest.fixture
def ledger(tmp_path):
    return build_ledger(tmp_path / "ledger.xlsx")


@pytest.fixture
def appended(ledger, docs, tmp_path):
    sheet, tables = choose_sheet(ledger)
    profile = suggest_profile(tables[sheet])
    records, flags = extract_records(docs["statement_dc"])
    result = append_records(
        ledger, records, profile, output_path=tmp_path / "out.xlsx", flags=flags
    )
    return result, load_workbook(result.output)


def test_rows_land_directly_under_the_existing_data(appended):
    result, wb = appended
    assert result.first_new_row == 12
    table = describe_sheet(wb["Transactions"])
    assert table.first_data_row == 4
    assert table.last_data_row == 19


def test_existing_rows_are_untouched(ledger, appended):
    _, wb = appended
    before = load_workbook(ledger)["Transactions"]
    after = wb["Transactions"]
    for row in range(1, 12):
        for col in range(1, 8):
            assert before.cell(row=row, column=col).value == after.cell(row=row, column=col).value


def test_the_totals_row_grows_to_cover_the_new_rows(appended):
    _, wb = appended
    ws = wb["Transactions"]
    table = describe_sheet(ws)
    totals_row = table.total_rows[0]
    for column in ("D", "E", "F"):
        formula = ws[f"{column}{totals_row}"].value
        assert range_in(formula) == f"{column}{table.first_data_row}:{column}{table.last_data_row}"


def test_the_total_actually_adds_up_to_the_new_figure(appended):
    result, wb = appended
    ws = wb["Transactions"]
    table = describe_sheet(ws)
    deposits = sum_of(ws, range_in(ws[f"D{table.total_rows[0]}"].value))
    assert deposits == Decimal("828") + result.money_in


def test_a_summary_sheet_pointing_here_is_repointed_too(appended):
    _, wb = appended
    assert wb["Summary"]["B1"].value == "=SUM(Transactions!D4:D19)"


def test_per_row_formulas_are_filled_down(appended):
    _, wb = appended
    ws = wb["Transactions"]
    for row in range(12, 20):
        assert ws.cell(row=row, column=6).value == f"=D{row}-E{row}"


def test_amounts_are_written_as_numbers_not_text(appended):
    _, wb = appended
    ws = wb["Transactions"]
    written = [ws.cell(row=r, column=c).value for r in range(12, 20) for c in (4, 5)]
    assert any(isinstance(v, float) for v in written)
    assert not any(isinstance(v, str) for v in written)


def test_new_rows_inherit_the_formatting_of_the_rows_above(appended):
    _, wb = appended
    ws = wb["Transactions"]
    assert ws["D12"].number_format == ws["D11"].number_format
    assert ws["A12"].number_format == ws["A11"].number_format


def test_the_autofilter_covers_the_new_rows(appended):
    _, wb = appended
    assert wb["Transactions"].auto_filter.ref == "A3:G19"


def test_running_the_same_documents_again_adds_nothing(appended, docs, tmp_path):
    result, _ = appended
    sheet, tables = choose_sheet(result.output)
    profile = suggest_profile(tables[sheet])
    records, _ = extract_records(docs["statement_dc"])
    second = append_records(result.output, records, profile, output_path=tmp_path / "again.xlsx")
    assert second.added == []
    assert len(second.skipped_duplicates) == 8


def test_dedupe_works_on_a_book_that_has_no_tracking_sheet(ledger, docs, tmp_path):
    """A hand-kept book with rows already in it must not gain copies of them."""
    sheet, tables = choose_sheet(ledger)
    profile = suggest_profile(tables[sheet])
    records, _ = extract_records(docs["statement_dc"])
    first = append_records(ledger, records, profile, output_path=tmp_path / "a.xlsx", track_state=False)
    second = append_records(first.output, records, profile, output_path=tmp_path / "b.xlsx", track_state=False)
    assert len(first.added) == 8
    assert second.added == []


def test_an_excel_table_range_is_extended(tmp_path, docs):
    book = build_ledger(tmp_path / "tbl.xlsx", with_table=True)
    sheet, tables = choose_sheet(book)
    records, _ = extract_records(docs["statement_dc"])
    result = append_records(book, records, suggest_profile(tables[sheet]), output_path=tmp_path / "tbl_out.xlsx")
    ws = load_workbook(result.output)["Transactions"]
    assert list(ws.tables.values())[0].ref == "A3:G19"


def test_invoices_go_to_an_invoice_register(tmp_path, docs):
    book = build_invoice_book(tmp_path / "inv.xlsx")
    sheet, tables = choose_sheet(book)
    profile = suggest_profile(tables[sheet])
    assert profile.kinds == ["invoice"]
    records, flags = extract_records(docs["invoice"])
    result = append_records(book, records, profile, output_path=tmp_path / "inv_out.xlsx", flags=flags)
    ws = load_workbook(result.output)["Invoices"]
    assert ws["B6"].value == "Brightline Media Ltd"
    assert ws["E6"].value == 4750.0
    assert ws["G6"].value == "=E6+F6"
    assert result.invoice_total == Decimal("5700.00")


def test_records_for_another_kind_of_sheet_are_reported_not_dropped(ledger, docs, tmp_path):
    sheet, tables = choose_sheet(ledger)
    profile = suggest_profile(tables[sheet])
    records, _ = extract_records(docs["invoice"])
    result = append_records(ledger, records, profile, output_path=tmp_path / "x.xlsx")
    assert result.added == []
    assert len(result.skipped_other_kind) == 1


def test_flags_are_written_to_a_review_notes_sheet(ledger, docs, tmp_path):
    from ledgerflow.models import Flag

    sheet, tables = choose_sheet(ledger)
    profile = suggest_profile(tables[sheet])
    records, _ = extract_records(docs["statement_dc"])
    flags = [Flag("statement.pdf", "page 1", "amount", "1,2O0.00", "Digit could be a letter O.")]
    result = append_records(ledger, records, profile, output_path=tmp_path / "flag.xlsx", flags=flags)
    ws = load_workbook(result.output)["Review Notes"]
    assert ws["B2"].value == "statement.pdf"
    assert "letter O" in ws["F2"].value


def test_a_dry_run_writes_nothing(ledger, docs, tmp_path):
    sheet, tables = choose_sheet(ledger)
    profile = suggest_profile(tables[sheet])
    records, _ = extract_records(docs["statement_dc"])
    out = tmp_path / "never.xlsx"
    result = append_records(ledger, records, profile, output_path=out, dry_run=True)
    assert len(result.added) == 8
    assert not out.exists()
