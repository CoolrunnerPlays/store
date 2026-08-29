"""Regressions from the audit run: a fully populated workbook must not be touched.

The audit uploaded a spreadsheet that already contained every transaction in the
documents. The expected result was nothing appended and no change to the file.
What happened instead was a full second copy of the ledger. These tests pin down
each cause so it cannot come back.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from fixtures import build_populated_ledger
from ledgerflow.append import (
    NO_NEW_MESSAGE,
    PreservationError,
    append_records,
    choose_sheet,
    plan_append,
)
from ledgerflow.extract import extract_records
from ledgerflow.mapping import suggest_profile
from ledgerflow.models import Record

VARIANTS = ["identical", "reworded", "textdates", "signflip"]


@pytest.fixture
def statement_records(docs):
    records = []
    for key in ("statement_dc", "statement_amount"):
        records.extend(extract_records(docs[key])[0])
    assert len(records) == 12
    return records


def populated(tmp_path, records, variant):
    path = build_populated_ledger(tmp_path / f"full_{variant}.xlsx", records, variant=variant)
    sheet, tables = choose_sheet(path)
    return path, suggest_profile(tables[sheet])


@pytest.mark.parametrize("variant", VARIANTS)
def test_a_fully_populated_book_gains_nothing(tmp_path, statement_records, variant):
    book, profile = populated(tmp_path, statement_records, variant)
    result = append_records(book, statement_records, profile, output_path=tmp_path / "out.xlsx")
    assert result.added == []
    assert result.status == "no_new_records"
    assert result.message == NO_NEW_MESSAGE


@pytest.mark.parametrize("variant", VARIANTS)
def test_no_file_is_written_when_there_is_nothing_new(tmp_path, statement_records, variant):
    """The mandatory pre-commit check: abort before the workbook is opened to write."""
    book, profile = populated(tmp_path, statement_records, variant)
    output = tmp_path / "should_not_exist.xlsx"
    before = book.read_bytes()

    result = append_records(book, statement_records, profile, output_path=output)

    assert not output.exists()
    assert result.file_written is False
    assert result.output is None
    assert book.read_bytes() == before, "the source workbook must be byte-identical"


@pytest.mark.parametrize("variant", VARIANTS)
def test_every_record_is_accounted_for_as_a_duplicate(tmp_path, statement_records, variant):
    book, profile = populated(tmp_path, statement_records, variant)
    result = plan_append(book, statement_records, profile)
    assert len(result.skipped_duplicates) == len(statement_records)
    assert result.unmatchable == []


def test_one_new_row_among_a_full_book_still_gets_through(tmp_path, statement_records):
    """Suppressing duplicates must not suppress genuine new work."""
    book, profile = populated(tmp_path, statement_records, "textdates")
    fresh = Record(
        kind="transaction", source="week33.pdf", date=date(2026, 5, 4),
        description="NEW SUPPLIER ONBOARDING", withdrawal=Decimal("777.77"), amount=Decimal("-777.77"),
    )
    result = append_records(book, statement_records + [fresh], profile, output_path=tmp_path / "out.xlsx")
    assert [r.description for r in result.added] == ["NEW SUPPLIER ONBOARDING"]
    assert result.status == "written"
    assert result.verified


def test_the_run_aborts_when_existing_rows_cannot_be_read(tmp_path, statement_records):
    """Fail closed: unable to compare means unable to append safely."""
    book, profile = populated(tmp_path, statement_records, "identical")
    # Take away the columns the matcher needs, as a bad mapping would.
    profile.columns = {letter: field for letter, field in profile.columns.items()
                       if field not in {"amount", "deposit", "withdrawal", "total", "date"}}
    result = plan_append(book, statement_records, profile)
    assert result.status == "unsafe_to_match"
    assert result.added == []


def test_running_twice_over_a_book_it_just_filled_adds_nothing(tmp_path, statement_records):
    """The weekly routine: last week's output is this week's input."""
    book = build_populated_ledger(tmp_path / "empty_start.xlsx", statement_records[:2], variant="identical")
    sheet, tables = choose_sheet(book)
    profile = suggest_profile(tables[sheet])

    first = append_records(book, statement_records, profile, output_path=tmp_path / "run1.xlsx")
    assert len(first.added) == 10

    sheet, tables = choose_sheet(first.output)
    second = append_records(first.output, statement_records, suggest_profile(tables[sheet]),
                            output_path=tmp_path / "run2.xlsx")
    assert second.added == []
    assert second.status == "no_new_records"
    assert not (tmp_path / "run2.xlsx").exists()


# --- formatting preservation -------------------------------------------------


def test_row_heights_below_the_insert_move_with_their_rows(tmp_path, statement_records):
    """openpyxl leaves row_dimensions behind on insert, mangling the sheet below."""
    book, profile = populated(tmp_path, statement_records, "textdates")
    before = load_workbook(book)["Transactions"]
    totals_row = before.max_row
    totals_height = before.row_dimensions[totals_row].height
    assert totals_height == 34

    fresh = Record(kind="transaction", source="w.pdf", date=date(2026, 5, 4),
                   description="BRAND NEW", withdrawal=Decimal("777.77"), amount=Decimal("-777.77"))
    result = append_records(book, [fresh], profile, output_path=tmp_path / "out.xlsx")

    after = load_workbook(result.output)["Transactions"]
    assert after.row_dimensions[totals_row + 1].height == totals_height
    assert after.cell(row=totals_row + 1, column=2).value == "TOTAL"


def test_every_style_facet_of_existing_cells_is_identical(tmp_path, statement_records):
    book, profile = populated(tmp_path, statement_records, "textdates")
    fresh = Record(kind="transaction", source="w.pdf", date=date(2026, 5, 4),
                   description="BRAND NEW", withdrawal=Decimal("777.77"), amount=Decimal("-777.77"))
    result = append_records(book, [fresh], profile, output_path=tmp_path / "out.xlsx")

    before = load_workbook(book)["Transactions"]
    after = load_workbook(result.output)["Transactions"]
    for row in range(1, before.max_row - 1):
        for col in range(1, 8):
            old, new = before.cell(row=row, column=col), after.cell(row=row, column=col)
            assert old.font.name == new.font.name
            assert old.font.sz == new.font.sz
            assert old.font.b == new.font.b
            assert old.font.color.rgb if old.font.color else None == (new.font.color.rgb if new.font.color else None)
            assert old.fill.fgColor.rgb == new.fill.fgColor.rgb
            assert old.border.bottom.style == new.border.bottom.style
            assert old.alignment.horizontal == new.alignment.horizontal
            assert old.number_format == new.number_format


def test_new_rows_inherit_the_font_of_the_row_above(tmp_path, statement_records):
    book, profile = populated(tmp_path, statement_records, "textdates")
    fresh = Record(kind="transaction", source="w.pdf", date=date(2026, 5, 4),
                   description="BRAND NEW", withdrawal=Decimal("777.77"), amount=Decimal("-777.77"))
    result = append_records(book, [fresh], profile, output_path=tmp_path / "out.xlsx")

    ws = load_workbook(result.output)["Transactions"]
    template_row, new_row = 15, 16
    assert ws.cell(row=new_row, column=2).font.name == ws.cell(row=template_row, column=2).font.name
    assert ws.cell(row=new_row, column=2).font.sz == ws.cell(row=template_row, column=2).font.sz
    assert ws.row_dimensions[new_row].height == ws.row_dimensions[template_row].height


def test_the_verifier_refuses_a_file_whose_existing_cells_changed(tmp_path, statement_records, monkeypatch):
    """A guarantee that is never checked is only a hope; prove the check bites."""
    import ledgerflow.append as append_module

    book, profile = populated(tmp_path, statement_records, "textdates")
    fresh = Record(kind="transaction", source="w.pdf", date=date(2026, 5, 4),
                   description="BRAND NEW", withdrawal=Decimal("777.77"), amount=Decimal("-777.77"))

    def damage(ws, source_row, target_row, table):
        from openpyxl.styles import Font
        ws.cell(row=1, column=1).font = Font(name="Comic Sans MS", size=8)

    monkeypatch.setattr(append_module, "_copy_row_style", damage)
    with pytest.raises(PreservationError):
        append_records(book, [fresh], profile, output_path=tmp_path / "out.xlsx")


def test_the_verifier_catches_the_row_height_bug(tmp_path, statement_records, monkeypatch):
    import ledgerflow.append as append_module

    book, profile = populated(tmp_path, statement_records, "textdates")
    fresh = Record(kind="transaction", source="w.pdf", date=date(2026, 5, 4),
                   description="BRAND NEW", withdrawal=Decimal("777.77"), amount=Decimal("-777.77"))

    monkeypatch.setattr(append_module, "_shift_row_dimensions", lambda ws, at_row, count: None)
    with pytest.raises(PreservationError, match="height"):
        append_records(book, [fresh], profile, output_path=tmp_path / "out.xlsx")
