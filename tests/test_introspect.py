from openpyxl import Workbook, load_workbook

from fixtures import build_invoice_book, build_ledger
from ledgerflow.introspect import describe_sheet, describe_workbook, find_header_row
from ledgerflow.mapping import guess_field, suggest_profile


def test_header_row_is_found_below_a_title(tmp_path):
    ws = load_workbook(build_ledger(tmp_path / "l.xlsx"))["Transactions"]
    assert find_header_row(ws) == 3


def test_data_block_stops_above_the_totals_row(tmp_path):
    table = describe_sheet(load_workbook(build_ledger(tmp_path / "l.xlsx"))["Transactions"])
    assert (table.first_data_row, table.last_data_row) == (4, 11)
    assert table.total_rows == [13]


def test_a_computed_column_is_recognised(tmp_path):
    table = describe_sheet(load_workbook(build_ledger(tmp_path / "l.xlsx"))["Transactions"])
    net = table.column_by_header("Net")
    assert net.is_computed
    assert net.formula_template == "=D11-E11"


def test_a_summary_block_does_not_pass_as_a_ledger(tmp_path):
    tables = describe_workbook(load_workbook(build_ledger(tmp_path / "l.xlsx")))
    assert tables["Transactions"].looks_like_table
    assert not tables["Summary"].looks_like_table


def test_a_sheet_with_no_table_is_skipped(tmp_path):
    wb = Workbook()
    wb.active["A1"] = "just a note"
    path = tmp_path / "empty.xlsx"
    wb.save(path)
    assert describe_workbook(load_workbook(path)) == {}


def test_headers_map_to_fields():
    assert guess_field("Withdrawal") == "withdrawal"
    assert guess_field("Money In") == "deposit"
    assert guess_field("Invoice No") == "reference"
    assert guess_field("VAT") == "tax"
    assert guess_field("Widget colour") is None


def test_computed_columns_are_never_mapped(tmp_path):
    table = describe_sheet(load_workbook(build_ledger(tmp_path / "l.xlsx"))["Transactions"])
    profile = suggest_profile(table)
    assert "F" not in profile.columns


def test_a_sheet_of_invoice_columns_is_marked_for_invoices(tmp_path):
    table = describe_sheet(load_workbook(build_invoice_book(tmp_path / "i.xlsx"))["Invoices"])
    assert suggest_profile(table).kinds == ["invoice"]


def test_a_profile_survives_a_round_trip(tmp_path):
    from ledgerflow.mapping import Profile

    table = describe_sheet(load_workbook(build_ledger(tmp_path / "l.xlsx"))["Transactions"])
    profile = suggest_profile(table)
    path = profile.save(tmp_path / "p.json")
    assert Profile.load(path) == profile
